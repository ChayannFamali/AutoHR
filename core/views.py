import json
import logging

import openpyxl
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Avg, Count, Q
from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_http_methods, require_POST
from django.views.generic import CreateView, DetailView, ListView, UpdateView
from openpyxl.styles import Alignment, Font, PatternFill

from accounts.decorators import hr_required
from notifications.services import NotificationService
from resume.models import Resume

from .forms import ApplicationForm, CompanyReviewForm, JobCreateForm
from .models import (Application, Candidate, Company, CompanyReview, Job,
                     SavedJob, SavedSearch)
from .ratelimit import rate_limit
from .utils import (can_export_data, can_manage_applications,
                    can_view_sensitive_data)

logger = logging.getLogger(__name__)


@login_required
@require_http_methods(["POST"])
def add_note_to_application(request, application_id):
    # Только HR и админы могут добавлять заметки
    if not can_manage_applications(request.user):
        return JsonResponse({
            'success': False,
            'message': 'У вас нет прав для добавления заметок'
        })
    try:
        application = get_object_or_404(Application, id=application_id)
        
        data = json.loads(request.body)
        note_text = data.get('note', '').strip()
        
        if not note_text:
            return JsonResponse({
                'success': False,
                'message': 'Заметка не может быть пустой'
            })
        
        # Добавляем заметку к существующим заметкам
        current_notes = application.notes or ""
        timestamp = timezone.now().strftime("%d.%m.%Y %H:%M")
        new_note = f"[{timestamp}] {request.user.get_full_name() or request.user.username}: {note_text}"
        
        if current_notes:
            application.notes = f"{current_notes}\n\n{new_note}"
        else:
            application.notes = new_note
            
        application.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Заметка успешно добавлена',
            'notes': application.notes
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Ошибка при добавлении заметки: {str(e)}'
        })
        
        
@hr_required
@require_POST
def update_application_status(request, application_id):
    """Обновление статуса заявки"""
    try:
        application = get_object_or_404(Application, id=application_id)
        data = json.loads(request.body)
        new_status = data.get('status')

        if new_status in ['approved', 'rejected']:
            application.status = new_status
            application.save()

            status_text = 'одобрена' if new_status == 'approved' else 'отклонена'

            return JsonResponse({
                'success': True,
                'message': f'Заявка {status_text}',
                'new_status': application.get_status_display()
            })
        else:
            return JsonResponse({
                'success': False,
                'message': 'Неверный статус'
            })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Ошибка: {str(e)}'
        })


@hr_required
@require_http_methods(["POST"])
@rate_limit(key='update_app_status', limit=30, period=60)
def update_application_status_htmx(request, application_id):
    """HTMX-эндпоинт: обновление статуса заявки, возвращает partial-фрагмент строки.

    Используется из templates/core/application_list.html через hx-post + hx-swap=outerHTML.
    При успехе заменяет строку таблицы на обновлённую.
    """
    application = get_object_or_404(Application, id=application_id)
    new_status = request.POST.get('status')

    if new_status not in ['approved', 'rejected']:
        from django.http import HttpResponse
        return HttpResponse(
            '<div class="alert alert-danger">Неверный статус</div>',
            status=400,
        )

    application.status = new_status
    application.save()

    return render(request, 'core/partials/application_row.html', {
        'application': application,
        'can_manage_applications': can_manage_applications(request.user),
    })


@hr_required
@require_http_methods(["GET"])
def application_detail_htmx(request, application_id):
    """HTMX-эндпоинт: возвращает HTML-partial модального окна деталей заявки.

    Используется из templates/core/application_list.html через hx-get + data-bs-toggle=modal.
    """
    application = get_object_or_404(Application, id=application_id)

    if request.user.is_candidate():
        try:
            candidate = Candidate.objects.get(user=request.user)
            if application.candidate != candidate:
                raise Http404("Заявка не найдена")
        except Candidate.DoesNotExist:
            raise Http404("Заявка не найдена")

    response = render(request, 'core/application_detail_modal.html', {
        'application': application,
        'can_view_sensitive_data': can_view_sensitive_data(request.user),
        'can_manage_applications': can_manage_applications(request.user),
    })
    response['HX-Trigger'] = json.dumps({'showModal': 'detailModal'})
    return response


@hr_required
@require_http_methods(["POST"])
def schedule_interview_htmx(request, application_id):
    """HTMX-эндпоинт: создаёт собеседование и возвращает обновлённую строку.

    Принимает: date, time, format, location.
    """
    if not can_manage_applications(request.user):
        return HttpResponse(
            '<div class="alert alert-danger">Нет прав</div>',
            status=403,
        )

    application = get_object_or_404(Application, id=application_id)

    date_str = request.POST.get('date')
    time_str = request.POST.get('time')
    format_type = request.POST.get('format')
    location = request.POST.get('location', '')

    if not all([date_str, time_str, format_type]):
        return HttpResponse(
            '<div class="alert alert-danger">Заполните все обязательные поля</div>',
            status=400,
        )

    from datetime import datetime, timezone
    try:
        scheduled_datetime = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M")
        scheduled_datetime = scheduled_datetime.replace(tzinfo=timezone.utc)
    except ValueError:
        return HttpResponse(
            '<div class="alert alert-danger">Неверный формат даты или времени</div>',
            status=400,
        )

    if scheduled_datetime <= datetime.now(timezone.utc):
        return HttpResponse(
            '<div class="alert alert-danger">Дата должна быть в будущем</div>',
            status=400,
        )

    from calendar_app.models import Interview, InterviewType

    interview_type, _ = InterviewType.objects.get_or_create(
        name='Стандартное собеседование',
        defaults={
            'duration_minutes': 60,
            'is_active': True,
            'description': 'Стандартное собеседование с кандидатом',
        },
    )

    Interview.objects.create(
        application=application,
        candidate=application.candidate,
        interviewer=request.user,
        interview_type=interview_type,
        scheduled_at=scheduled_datetime,
        format=format_type,
        location=location,
        status='scheduled',
        duration_minutes=interview_type.duration_minutes,
    )

    application.status = 'approved'
    application.save()

    response = render(request, 'core/partials/application_row.html', {
        'application': application,
        'can_manage_applications': can_manage_applications(request.user),
    })
    response['HX-Trigger'] = json.dumps({
        'showToast': {'message': 'Собеседование запланировано', 'type': 'success'},
    })
    return response


@hr_required
@require_http_methods(["POST"])
def add_note_htmx(request, application_id):
    """HTMX-эндпоинт: добавляет заметку к заявке.

    Принимает: note (textarea content).
    Возвращает: обновлённый блок заметок.
    """
    if not can_manage_applications(request.user):
        return HttpResponse(
            '<div class="alert alert-danger">Нет прав</div>',
            status=403,
        )

    application = get_object_or_404(Application, id=application_id)
    note_text = (request.POST.get('note') or '').strip()

    if not note_text:
        return HttpResponse(
            '<div class="alert alert-danger">Заметка не может быть пустой</div>',
            status=400,
        )

    current_notes = application.notes or ''
    timestamp = timezone.now().strftime('%d.%m.%Y %H:%M')
    new_note = f'[{timestamp}] {request.user.get_full_name() or request.user.username}: {note_text}'

    if current_notes:
        application.notes = f'{current_notes}\n\n{new_note}'
    else:
        application.notes = new_note
    application.save()

    response = render(request, 'core/partials/notes_block.html', {
        'application': application,
        'can_view_sensitive_data': can_view_sensitive_data(request.user),
    })
    response['HX-Trigger'] = json.dumps({
        'showToast': {'message': 'Заметка добавлена', 'type': 'success'},
    })
    return response


@login_required
@require_POST
def delete_job_htmx(request, job_id):
    """HTMX-эндпоинт: удаляет вакансию (только владельца).

    При успехе возвращает 200 + OOB-команду удалить карточку из DOM.
    """
    job = get_object_or_404(Job, id=job_id, created_by=request.user)
    job.delete()
    response = HttpResponse(status=200)
    response['HX-Trigger'] = json.dumps({
        'showToast': {'message': 'Вакансия удалена', 'type': 'success'},
    })
    return response


@hr_required
@require_POST
def update_interview_status_htmx(request, interview_id):
    """HTMX-эндпоинт: обновление статуса собеседования → возвращает обновлённую карточку."""
    from calendar_app.models import Interview

    if not can_manage_applications(request.user):
        return HttpResponse('<div class="alert alert-danger">Нет прав</div>', status=403)

    interview = get_object_or_404(Interview, id=interview_id)
    new_status = request.POST.get('status')

    if new_status not in ['confirmed', 'completed', 'cancelled']:
        return HttpResponse('<div class="alert alert-danger">Неверный статус</div>', status=400)

    interview.status = new_status
    interview.save()

    if new_status == 'completed':
        interview.application.status = 'interviewed'
        interview.application.save()
    elif new_status == 'cancelled':
        interview.application.status = 'pending'
        interview.application.save()

    response = render(request, 'calendar_app/partials/interview_card.html', {
        'interview': interview,
        'can_manage_applications': True,
    })
    response['HX-Trigger'] = json.dumps({
        'showToast': {'message': f'Статус: {interview.get_status_display()}', 'type': 'success'},
    })
    return response


@hr_required
@require_POST
def reschedule_interview_htmx(request, interview_id):
    """HTMX-эндпоинт: перенос собеседования → возвращает обновлённую карточку."""
    from datetime import datetime, timezone
    from calendar_app.models import Interview

    if not can_manage_applications(request.user):
        return HttpResponse('<div class="alert alert-danger">Нет прав</div>', status=403)

    interview = get_object_or_404(Interview, id=interview_id)

    date_str = request.POST.get('date')
    time_str = request.POST.get('time')
    reason = request.POST.get('reason', '')

    if not date_str or not time_str:
        return HttpResponse(
            '<div class="alert alert-danger">Укажите новые дату и время</div>',
            status=400,
        )

    try:
        new_datetime = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M")
        new_datetime = new_datetime.replace(tzinfo=timezone.utc)
    except ValueError:
        return HttpResponse(
            '<div class="alert alert-danger">Неверный формат даты или времени</div>',
            status=400,
        )

    if new_datetime <= datetime.now(timezone.utc):
        return HttpResponse(
            '<div class="alert alert-danger">Новая дата должна быть в будущем</div>',
            status=400,
        )

    interview.scheduled_at = new_datetime
    interview.status = 'rescheduled'
    if reason:
        interview.notes = f"{interview.notes or ''}\n[Перенос] Причина: {reason}"
    interview.save()

    response = render(request, 'calendar_app/partials/interview_card.html', {
        'interview': interview,
        'can_manage_applications': True,
    })
    response['HX-Trigger'] = json.dumps({
        'showToast': {
            'message': f'Перенесено на {new_datetime.strftime("%d.%m.%Y в %H:%M")}',
            'type': 'success',
        },
    })
    return response


@hr_required
@require_POST
def save_interview_feedback_htmx(request, interview_id):
    """HTMX-эндпоинт: сохранить feedback собеседования → возвращает обновлённую карточку."""
    from calendar_app.models import Interview

    if not can_manage_applications(request.user):
        return HttpResponse('<div class="alert alert-danger">Нет прав</div>', status=403)

    interview = get_object_or_404(Interview, id=interview_id)
    interview.feedback = request.POST.get('feedback', '')
    rating = request.POST.get('rating')
    if rating:
        try:
            interview.rating = int(rating)
        except (TypeError, ValueError):
            pass
    interview.save()

    response = render(request, 'calendar_app/partials/interview_card.html', {
        'interview': interview,
        'can_manage_applications': True,
    })
    response['HX-Trigger'] = json.dumps({
        'showToast': {'message': 'Отзыв сохранён', 'type': 'success'},
    })
    return response

@hr_required
@login_required
@require_http_methods(["POST"])
def schedule_interview_for_application(request, application_id):
    """Быстрое планирование собеседования"""
    
    if not can_manage_applications(request.user):
        return JsonResponse({
            'success': False,
            'message': 'У вас нет прав для планирования собеседований'
        })
    
    try:
        application = get_object_or_404(Application, id=application_id)
        
        date_str = request.POST.get('date')
        time_str = request.POST.get('time')
        format_type = request.POST.get('format')
        location = request.POST.get('location', '')
        
        if not all([date_str, time_str, format_type]):
            return JsonResponse({
                'success': False,
                'message': 'Заполните все обязательные поля (дата, время, формат)'
            })
        
        from datetime import datetime, timezone
        try:
            scheduled_datetime = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M")
            scheduled_datetime = scheduled_datetime.replace(tzinfo=timezone.utc)
        except ValueError:
            return JsonResponse({
                'success': False,
                'message': 'Неверный формат даты или времени'
            })
        
        # Проверяем, что дата в будущем
        if scheduled_datetime <= datetime.now(timezone.utc):
            return JsonResponse({
                'success': False,
                'message': 'Дата собеседования должна быть в будущем'
            })
        
        from calendar_app.models import Interview, InterviewType

        # Получаем или создаем дефолтный тип собеседования
        interview_type, created = InterviewType.objects.get_or_create(
            name='Стандартное собеседование',
            defaults={
                'duration_minutes': 60,
                'is_active': True,
                'description': 'Стандартное собеседование с кандидатом'
            }
        )
        
        if created:
            print(f"Создан новый тип собеседования: {interview_type.name}")
        
        interview = Interview.objects.create(
            application=application,
            candidate=application.candidate,
            interviewer=request.user,
            interview_type=interview_type,
            scheduled_at=scheduled_datetime,
            format=format_type,
            location=location,
            status='scheduled',
            duration_minutes=interview_type.duration_minutes
        )
        
        application.status = 'approved'
        application.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Собеседование запланировано на {scheduled_datetime.strftime("%d.%m.%Y в %H:%M")}'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Ошибка: {str(e)}'
        })



@login_required
@hr_required
def export_applications_excel(request):
    # Создаем новую книгу Excel
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Заявки кандидатов'
    
    # Заголовки столбцов
    headers = [
        'ID', 'Кандидат', 'Email', 'Телефон', 'Вакансия', 'Компания', 
        'AI Оценка', 'Статус', 'Дата подачи', 'Сопроводительное письмо'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    
    applications = Application.objects.select_related(
        'candidate', 'job', 'job__company'
    ).order_by('-applied_at')
    
    for row_num, application in enumerate(applications, 2):
        worksheet.cell(row=row_num, column=1, value=application.id)
        worksheet.cell(row=row_num, column=2, value=application.candidate.full_name)
        worksheet.cell(row=row_num, column=3, value=application.candidate.email)
        worksheet.cell(row=row_num, column=4, value=application.candidate.phone or '-')
        worksheet.cell(row=row_num, column=5, value=application.job.title)
        worksheet.cell(row=row_num, column=6, value=application.job.company.name)
        worksheet.cell(row=row_num, column=7, value=application.ai_score or '-')
        worksheet.cell(row=row_num, column=8, value=application.get_status_display())
        worksheet.cell(row=row_num, column=9, value=application.applied_at.strftime('%d.%m.%Y %H:%M'))
        worksheet.cell(row=row_num, column=10, value=application.cover_letter[:100] + '...' if len(application.cover_letter) > 100 else application.cover_letter)
    
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Создаем HTTP ответ с Excel файлом
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="applications_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    workbook.save(response)
    return response

@hr_required
def create_job(request):
    """Создание новой вакансии"""
    if request.method == 'POST':
        form = JobCreateForm(request.POST)
        if form.is_valid():
            job = form.save(commit=False)
            job.created_by = request.user
            job.save()
            
            messages.success(request, f'Вакансия "{job.title}" успешно создана!')
            return redirect('core:job_detail', pk=job.id)
        else:
            messages.error(request, 'Пожалуйста, исправьте ошибки в форме.')
    else:
        form = JobCreateForm()
    
    return render(request, 'core/create_job.html', {'form': form})

def _filter_jobs(request, base_qs):
    """Единая логика фильтрации вакансий (Этап 3.3).

    Поддерживает GET-параметры:
      - search: полнотекстовый поиск (icontains; в Postgres — SearchVector)
      - experience_level: junior/middle/senior
      - remote_work: true/false
      - location: подстрока локации
      - employment_type: full_time/part_time/internship/project/remote
      - salary_min, salary_max: диапазон зарплаты (пересечение с вилкой)
      - skills: список навыков (multi-select, любой совпадает)
      - sort: date / salary_desc / salary_asc / ai_relevance
    """
    qs = base_qs.filter(status='active').select_related('company')

    search = request.GET.get('search', '').strip()
    if search:
        from django.db import connection
        if connection.vendor == 'postgresql':
            from django.contrib.postgres.search import (
                SearchQuery, SearchRank, SearchVector,
            )
            vector = (
                SearchVector('title', weight='A')
                + SearchVector('description', weight='B')
                + SearchVector('requirements', weight='C')
                + SearchVector('company__name', weight='B')
            )
            qs = qs.annotate(
                search=vector,
                rank=SearchRank(vector, SearchQuery(search)),
            ).filter(search=SearchQuery(search)).order_by('-rank', '-created_at')
        else:
            qs = qs.filter(
                Q(title__icontains=search)
                | Q(description__icontains=search)
                | Q(requirements__icontains=search)
                | Q(company__name__icontains=search)
            )

    experience_level = request.GET.get('experience_level')
    if experience_level:
        qs = qs.filter(experience_level=experience_level)

    remote_work = request.GET.get('remote_work')
    if remote_work:
        qs = qs.filter(remote_work=(remote_work == 'true'))

    location = request.GET.get('location', '').strip()
    if location:
        qs = qs.filter(location__icontains=location)

    employment_type = request.GET.get('employment_type')
    if employment_type:
        qs = qs.filter(employment_type=employment_type)

    salary_min = request.GET.get('salary_min')
    if salary_min and salary_min.isdigit():
        salary_min = int(salary_min)
        qs = qs.filter(
            Q(salary_max__gte=salary_min) | Q(salary_min__gte=salary_min)
        )

    salary_max = request.GET.get('salary_max')
    if salary_max and salary_max.isdigit():
        salary_max = int(salary_max)
        qs = qs.filter(
            Q(salary_min__lte=salary_max) | Q(salary_max__lte=salary_max)
        )

    skills_raw = request.GET.get('skills')
    if skills_raw:
        skill_list = [s.strip() for s in skills_raw.split(',') if s.strip()]
        if skill_list:
            from django.db.models import Q as Q2
            skill_q = Q2()
            for skill in skill_list:
                skill_q |= Q2(skills_required__icontains=f'"{skill}"')
            qs = qs.filter(skill_q)

    sort = request.GET.get('sort', 'date')
    if sort == 'salary_desc':
        qs = qs.order_by('-salary_max', '-salary_min', '-created_at')
    elif sort == 'salary_asc':
        qs = qs.order_by('salary_min', 'salary_max', '-created_at')
    elif sort == 'ai_relevance':
        if getattr(settings, 'AI_ENABLED', False):
            qs = qs.order_by('-views_count', '-created_at')
        else:
            qs = qs.order_by('-views_count', '-created_at')
    else:
        qs = qs.order_by('-created_at')

    return qs


def job_list(request):
    """Список всех активных вакансий с расширенными фильтрами."""
    jobs = _filter_jobs(request, Job.objects.all())

    saved_ids = (
        set(SavedJob.objects.filter(
            user=request.user, job__in=jobs
        ).values_list('job_id', flat=True))
        if request.user.is_authenticated else set()
    )

    context = {
        'jobs': jobs,
        'search_query': request.GET.get('search', ''),
        'selected_experience': request.GET.get('experience_level', ''),
        'selected_remote': request.GET.get('remote_work', ''),
        'selected_location': request.GET.get('location', ''),
        'selected_employment': request.GET.get('employment_type', ''),
        'selected_salary_min': request.GET.get('salary_min', ''),
        'selected_salary_max': request.GET.get('salary_max', ''),
        'selected_skills': request.GET.get('skills', ''),
        'selected_sort': request.GET.get('sort', 'date'),
        'saved_job_ids': saved_ids,
    }

    return render(request, 'core/job_list.html', context)


def job_list_htmx(request):
    """HTMX-эндпоинт: возвращает только список вакансий (без фильтров и обёртки).

    Использует _filter_jobs для единой логики фильтрации.
    """
    jobs = _filter_jobs(request, Job.objects.all())

    saved_ids = (
        set(SavedJob.objects.filter(
            user=request.user, job__in=jobs
        ).values_list('job_id', flat=True))
        if request.user.is_authenticated else set()
    )

    return render(request, 'core/partials/job_list.html', {
        'jobs': jobs,
        'saved_job_ids': saved_ids,
    })

@login_required
@require_http_methods(["POST"])
def delete_job(request, job_id):
    """Удаление вакансии"""
    job = get_object_or_404(Job, id=job_id, created_by=request.user)
    
    try:
        job_title = job.title
        job.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Вакансия "{job_title}" успешно удалена'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Ошибка при удалении: {str(e)}'
        })

@hr_required
def job_list_hr(request):
    """Список вакансий для HR"""
    jobs = Job.objects.filter(created_by=request.user).order_by('-created_at')
    return render(request, 'core/job_list_hr.html', {'jobs': jobs})

class JobListView(ListView):
    model = Job
    template_name = 'core/job_list.html'
    context_object_name = 'jobs'
    paginate_by = 10

    def get_queryset(self):
        return _filter_jobs(self.request, Job.objects.all())

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['saved_job_ids'] = (
            set(SavedJob.objects.filter(
                user=self.request.user, job__in=ctx['jobs']
            ).values_list('job_id', flat=True))
            if self.request.user.is_authenticated else set()
        )
        ctx['search_query'] = self.request.GET.get('search', '')
        ctx['selected_experience'] = self.request.GET.get('experience_level', '')
        ctx['selected_remote'] = self.request.GET.get('remote_work', '')
        ctx['selected_location'] = self.request.GET.get('location', '')
        ctx['selected_employment'] = self.request.GET.get('employment_type', '')
        ctx['selected_salary_min'] = self.request.GET.get('salary_min', '')
        ctx['selected_salary_max'] = self.request.GET.get('salary_max', '')
        ctx['selected_skills'] = self.request.GET.get('skills', '')
        ctx['selected_sort'] = self.request.GET.get('sort', 'date')
        return ctx

class JobDetailView(DetailView):
    model = Job
    template_name = 'core/job_detail.html'
    context_object_name = 'job'

    def get_queryset(self):
        return Job.objects.filter(status='active').select_related('company')

    def get_object(self, queryset=None):
        obj = super().get_object(queryset)
        obj.views_count += 1
        obj.save(update_fields=['views_count'])
        return obj

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['is_saved'] = (
            self.request.user.is_authenticated
            and SavedJob.objects.filter(user=self.request.user, job=self.object).exists()
        )
        from ai_analysis.services.recommendations import get_similar_jobs
        ctx['similar_jobs'] = get_similar_jobs(self.object.id, top_n=5)
        return ctx

@rate_limit(key='apply_job', limit=10, period=3600)
def apply_for_job(request, job_id):
    job = get_object_or_404(Job, id=job_id, status='active')
    
    if request.method == 'POST':
        form = ApplicationForm(request.POST)
        if form.is_valid():
            try:
                if request.user.is_authenticated and request.user.is_candidate():
                    candidate, created = Candidate.objects.get_or_create(
                        user=request.user,
                        defaults={
                            'first_name': form.cleaned_data['first_name'],
                            'last_name': form.cleaned_data['last_name'],
                            'email': form.cleaned_data['email'],
                            'phone': form.cleaned_data.get('phone', ''),
                        }
                    )
                    if not created:
                        candidate.first_name = form.cleaned_data['first_name']
                        candidate.last_name = form.cleaned_data['last_name']
                        candidate.email = form.cleaned_data['email']
                        candidate.phone = form.cleaned_data.get('phone', '')
                        candidate.save()
                else:
                    candidate, created = Candidate.objects.get_or_create(
                        email=form.cleaned_data['email'],
                        defaults={
                            'first_name': form.cleaned_data['first_name'],
                            'last_name': form.cleaned_data['last_name'],
                            'phone': form.cleaned_data.get('phone', ''),
                        }
                    )
                
                existing_application = Application.objects.filter(
                    candidate=candidate,
                    job=job
                ).first()
                
                if existing_application:
                    messages.warning(request, 'Вы уже подавали заявку на эту вакансию.')
                    return redirect('core:job_detail', pk=job.id)
                
                application = Application.objects.create(
                    candidate=candidate,
                    job=job,
                    cover_letter=form.cleaned_data.get('cover_letter', ''),
                    status='pending'
                )
                
                existing_resumes = Resume.objects.filter(
                    candidate=candidate,
                    status='processed'
                ).order_by('-uploaded_at')

                if existing_resumes.exists():
                    latest_resume = existing_resumes.first()

                    from django.conf import settings
                    if getattr(settings, 'AI_ENABLED', False):
                        from ai_analysis.tasks import match_candidate_with_job_task
                        match_candidate_with_job_task.delay(latest_resume.id, job.id)
                        messages.success(
                            request,
                            'Заявка отправлена! AI-оценка запущена в фоне.',
                        )
                    else:
                        messages.success(
                            request,
                            'Заявка отправлена! (AI-анализ выключен)',
                        )
                else:
                    messages.success(
                        request,
                        'Заявка отправлена! Загрузите резюме для скоринга.',
                    )

                try:
                    from notifications.services import NotificationService

                    # Уведомление кандидату
                    NotificationService.send_application_confirmation(application)
                    
                    hr_message = f'Новая заявка от {candidate.full_name} на вакансию "{job.title}"\n\n'
                    hr_message += f'Email кандидата: {candidate.email}\n'
                    hr_message += f'Телефон: {candidate.phone or "Не указан"}\n'
                    hr_message += f'Дата подачи: {application.applied_at.strftime("%d.%m.%Y %H:%M")}\n\n'
                    if application.cover_letter:
                        hr_message += f'Сопроводительное письмо:\n{application.cover_letter}'
                    NotificationService.send_hr_notification(
                        'Новая заявка',
                        hr_message,
                        job.created_by.email
                    )
                except Exception as e:
                    print(f"Ошибка отправки уведомлений: {e}")

                return redirect('resume:upload_resume_for_application', application_id=application.id)
                
            except Exception as e:
                messages.error(request, f'Произошла ошибка при отправке заявки: {str(e)}')
        else:
            messages.error(request, 'Пожалуйста, исправьте ошибки в форме.')
    else:
        form = ApplicationForm()
    
    return render(request, 'core/apply_job.html', {'job': job, 'form': form})

class ApplicationListView(LoginRequiredMixin, ListView):
    model = Application
    template_name = 'core/application_list.html'
    context_object_name = 'applications'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Application.objects.select_related('candidate', 'job', 'job__company').order_by('-applied_at')
        
        if self.request.user.is_candidate():
            try:
                candidate = Candidate.objects.get(user=self.request.user)
                queryset = queryset.filter(candidate=candidate)
            except Candidate.DoesNotExist:
                queryset = queryset.none()
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        base_queryset = self.get_queryset()
        
        total_applications = base_queryset.count()
        pending_applications = base_queryset.filter(status='pending').count()
        approved_applications = base_queryset.filter(status='approved').count()
        
        context.update({
            'can_view_sensitive_data': can_view_sensitive_data(self.request.user),
            'can_manage_applications': can_manage_applications(self.request.user),
            'can_export_data': can_export_data(self.request.user),
            'is_candidate': self.request.user.is_candidate(),
            'total_applications': total_applications,
            'pending_applications': pending_applications,
            'approved_applications': approved_applications,
        })
        
        return context

@login_required
def application_detail(request, application_id):
    application = get_object_or_404(Application, id=application_id)
    
    if request.user.is_candidate():
        try:
            candidate = Candidate.objects.get(user=request.user)
            if application.candidate != candidate:
                raise Http404("Заявка не найдена")
        except Candidate.DoesNotExist:
            raise Http404("Заявка не найдена")
    
    if request.method == 'GET':
        html = render_to_string('core/application_detail_modal.html', {
            'application': application,
            'can_view_sensitive_data': can_view_sensitive_data(request.user),
            'can_manage_applications': can_manage_applications(request.user),
        }, request=request)
        
        return JsonResponse({
            'success': True,
            'html': html
        })


class JobEditView(LoginRequiredMixin, UpdateView):
    model = Job
    form_class = JobCreateForm
    template_name = 'core/edit_job.html'
    
    def get_queryset(self):
        # Пользователь может редактировать только свои вакансии
        return Job.objects.filter(created_by=self.request.user)

    def get_success_url(self):
        return reverse('core:job_list_hr')


@login_required
@require_POST
def toggle_save_job(request, job_id):
    """HTMX-эндпоинт: добавить/убрать вакансию из избранного.

    Возвращает фрагмент кнопки «В избранном»/«В избранное»,
    который заменяется через hx-swap="outerHTML".
    """
    job = get_object_or_404(Job, id=job_id, status='active')

    saved, created = SavedJob.objects.get_or_create(
        user=request.user,
        job=job,
    )

    if not created:
        saved.delete()
        is_saved = False
    else:
        is_saved = True

    return render(request, 'core/partials/save_button.html', {
        'job': job,
        'is_saved': is_saved,
    })


@login_required
def favorites_list(request):
    """Страница «Избранное»: сохранённые вакансии и поиски."""
    saved_jobs = (
        SavedJob.objects
        .filter(user=request.user)
        .select_related('job', 'job__company')
        .order_by('-created_at')
    )
    saved_searches = (
        SavedSearch.objects
        .filter(user=request.user)
        .order_by('-last_used_at', '-created_at')
    )

    return render(request, 'core/favorites.html', {
        'saved_jobs': saved_jobs,
        'saved_searches': saved_searches,
    })


@login_required
@require_POST
def save_current_search(request):
    """Сохранить текущие GET-параметры как поиск пользователя."""
    name = (request.POST.get('name') or '').strip()
    if not name:
        return JsonResponse({'success': False, 'message': 'Укажите название поиска'})

    excluded = {'page'}
    params = {
        k: request.POST.get(k) or request.GET.get(k)
        for k in ('search', 'experience_level', 'remote_work',
                  'location', 'salary_min', 'salary_max',
                  'employment_type', 'skills', 'sort')
        if (request.POST.get(k) or request.GET.get(k))
    }
    params = {k: v for k, v in params.items() if k not in excluded and v}

    saved = SavedSearch.objects.create(
        user=request.user,
        name=name,
        params=params,
    )
    return JsonResponse({'success': True, 'id': saved.id, 'name': saved.name})


@login_required
def run_saved_search(request, search_id):
    """Применить сохранённый поиск: редирект на /jobs/? с параметрами."""
    saved = get_object_or_404(SavedSearch, id=search_id, user=request.user)

    saved.last_used_at = timezone.now()
    saved.save(update_fields=['last_used_at'])

    from django.http import QueryDict
    qs = QueryDict('', mutable=True)
    for k, v in saved.params.items():
        qs[k] = str(v)
    return redirect(f"{reverse('core:job_list')}?{qs.urlencode()}")


@login_required
@require_POST
def delete_saved_search(request, search_id):
    """Удалить сохранённый поиск (HTMX-эндпоинт)."""
    saved = get_object_or_404(SavedSearch, id=search_id, user=request.user)
    saved.delete()
    from django.http import HttpResponse
    return HttpResponse('', status=200)


def company_list(request):
    """Каталог компаний с поиском и фильтром по отрасли."""
    companies = Company.objects.annotate(
        _jobs_count=Count('job', filter=Q(job__status='active')),
        _avg_rating=Avg('reviews__rating'),
    ).order_by('name')

    search_query = request.GET.get('search')
    if search_query:
        companies = companies.filter(
            Q(name__icontains=search_query)
            | Q(description__icontains=search_query)
            | Q(industry__icontains=search_query)
        )

    industry = request.GET.get('industry')
    if industry:
        companies = companies.filter(industry=industry)

    industries = (
        Company.objects.exclude(industry='')
        .values_list('industry', flat=True)
        .distinct()
        .order_by('industry')
    )

    return render(request, 'core/company_list.html', {
        'companies': companies,
        'search_query': search_query,
        'selected_industry': industry,
        'industries': industries,
    })


def company_detail(request, slug):
    """Страница компании: hero, вакансии, отзывы."""
    company = get_object_or_404(Company, slug=slug)

    jobs = (
        company.job_set
        .filter(status='active')
        .select_related('company')
        .order_by('-created_at')
    )
    reviews = (
        company.reviews
        .select_related('author')
        .order_by('-created_at')[:20]
    )

    user_review = None
    if request.user.is_authenticated:
        user_review = company.reviews.filter(author=request.user).first()

    saved_job_ids = (
        set(SavedJob.objects.filter(
            user=request.user, job__in=jobs
        ).values_list('job_id', flat=True))
        if request.user.is_authenticated else set()
    )

    return render(request, 'core/company_detail.html', {
        'company': company,
        'jobs': jobs,
        'reviews': reviews,
        'user_review': user_review,
        'saved_job_ids': saved_job_ids,
    })


@login_required
@require_POST
def add_company_review(request, slug):
    """Добавить отзыв о компании."""
    company = get_object_or_404(Company, slug=slug)

    existing = CompanyReview.objects.filter(
        company=company, author=request.user
    ).first()
    if existing:
        messages.warning(
            request,
            'Вы уже оставляли отзыв о этой компании.',
        )
        return redirect('core:company_detail', slug=company.slug)

    form = CompanyReviewForm(request.POST)
    if form.is_valid():
        review = form.save(commit=False)
        review.company = company
        review.author = request.user
        review.save()
        messages.success(request, 'Спасибо! Отзыв опубликован.')
    else:
        messages.error(request, 'Пожалуйста, исправьте ошибки в форме.')

    return redirect('core:company_detail', slug=company.slug)