import csv
import io
import logging

import openpyxl
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics, pdfutils
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas

from ai_analysis.services.analysis_engine import AnalysisEngine
from core.models import Application, Candidate
from core.utils import (can_export_data, can_manage_applications,
                        can_view_sensitive_data)

from .forms import ResumeUploadForm
from .models import Candidate, Resume

logger = logging.getLogger(__name__)



@login_required
def upload_resume(request, application_id=None):
    application = None
    candidate = None
    
    if application_id:
        application = get_object_or_404(Application, id=application_id)
        candidate = application.candidate
    else:
        if request.user.is_candidate():
            try:
                candidate = Candidate.objects.get(user=request.user)
            except Candidate.DoesNotExist:
                candidate = Candidate.objects.create(
                    user=request.user,
                    first_name=request.user.first_name or 'Не указано',
                    last_name=request.user.last_name or 'Не указано',
                    email=request.user.email
                )
        else:
            messages.error(request, 'Только кандидаты могут загружать резюме.')
            return redirect('core:job_list')
    
    if request.method == 'POST':
        form = ResumeUploadForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                resume = form.save(commit=False)
                resume.candidate = candidate
                
                if application:
                    resume.application = application
                
                resume.save()
                
                try:
                    from ai_analysis.services.analysis_engine import \
                        AnalysisEngine
                    
                    resume.status = 'processing'
                    resume.save()
                    
                    analysis_engine = AnalysisEngine()
                    analysis_result = analysis_engine.analyze_resume(resume.id)
                    
                    if analysis_result['success']:
                        messages.success(request, 'Резюме успешно загружено и проанализировано!')
                        
                        if application:
                            match_result = analysis_engine.match_candidate_with_job(
                                resume.id, application.job.id
                            )
                            if match_result['success']:
                                application.ai_score = match_result['overall_score']
                                application.ai_feedback = match_result['recommendation']
                                application.save()
                    else:
                        messages.warning(request, 'Резюме загружено, но AI-анализ не удался.')
                        
                except ImportError:
                    resume.status = 'uploaded'
                    resume.save()
                    messages.success(request, 'Резюме успешно загружено! (AI-анализ недоступен)')
                    
                except Exception as e:
                    resume.status = 'error'
                    resume.processing_error = str(e)
                    resume.save()
                    messages.success(request, f'Резюме загружено! AI-анализ будет выполнен позже. ({str(e)[:100]})')
                
                if application:
                    return redirect('core:application_list')
                else:
                    return redirect('resume:resume_list')
                    
            except Exception as e:
                messages.error(request, f'Ошибка при загрузке резюме: {str(e)}')
        else:
            messages.error(request, 'Пожалуйста, исправьте ошибки в форме.')
    else:
        form = ResumeUploadForm()
    
    return render(request, 'resume/upload_resume.html', {
        'form': form,
        'application': application,
        'candidate': candidate,
    })

def reprocess_resume(request, resume_id):
    """Повторная обработка резюме"""
    if request.method == 'POST':
        try:
            resume = get_object_or_404(Resume, id=resume_id)
            
            analysis_engine = AnalysisEngine()
            resume.status = 'processing'
            resume.processing_error = ''
            resume.save()
            
            analysis_result = analysis_engine.analyze_resume(resume_id)
            
            if analysis_result['success']:
                messages.success(request, 'Резюме успешно переобработано!')
                
                if resume.application:
                    match_result = analysis_engine.match_candidate_with_job(
                        resume_id, 
                        resume.application.job.id
                    )
                    
                    if match_result['success']:
                        resume.application.ai_score = match_result['overall_score']
                        resume.application.ai_feedback = match_result['recommendation']
                        resume.application.save()
            else:
                messages.error(request, f'Ошибка при переобработке: {analysis_result.get("error")}')
        
        except Exception as e:
            messages.error(request, f'Ошибка: {str(e)}')
    
    return redirect('resume:resume_detail', pk=resume_id)



def export_resumes_excel(request):
    status = request.GET.get('status')
    language = request.GET.get('language')
    candidate = request.GET.get('candidate')
    
    resumes = Resume.objects.select_related('candidate').order_by('-uploaded_at')
    
    if status:
        resumes = resumes.filter(status=status)
    if language:
        resumes = resumes.filter(language=language)
    if candidate:
        resumes = resumes.filter(
            Q(candidate__first_name__icontains=candidate) |
            Q(candidate__last_name__icontains=candidate) |
            Q(candidate__email__icontains=candidate)
        )
    
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Резюме кандидатов'
    
    headers = [
        'ID', 'Кандидат', 'Email', 'Файл', 'Размер', 'Язык', 
        'Статус', 'Опыт (лет)', 'Навыки', 'Дата загрузки'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    
    for row_num, resume in enumerate(resumes, 2):
        worksheet.cell(row=row_num, column=1, value=resume.id)
        worksheet.cell(row=row_num, column=2, value=resume.candidate.full_name)
        worksheet.cell(row=row_num, column=3, value=resume.candidate.email)
        worksheet.cell(row=row_num, column=4, value=resume.original_filename)
        worksheet.cell(row=row_num, column=5, value=f"{resume.file_size / 1024:.1f} KB")
        worksheet.cell(row=row_num, column=6, value=resume.get_language_display())
        worksheet.cell(row=row_num, column=7, value=resume.get_status_display())
        worksheet.cell(row=row_num, column=8, value=resume.experience_years or '-')
        worksheet.cell(row=row_num, column=9, value=', '.join(resume.skills[:5]) if resume.skills else '-')
        worksheet.cell(row=row_num, column=10, value=resume.uploaded_at.strftime('%d.%m.%Y %H:%M'))
    
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="resumes_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    workbook.save(response)
    return response

def export_resumes_csv(request):
    status = request.GET.get('status')
    language = request.GET.get('language')
    candidate = request.GET.get('candidate')
    
    resumes = Resume.objects.select_related('candidate').order_by('-uploaded_at')
    
    if status:
        resumes = resumes.filter(status=status)
    if language:
        resumes = resumes.filter(language=language)
    if candidate:
        resumes = resumes.filter(
            Q(candidate__first_name__icontains=candidate) |
            Q(candidate__last_name__icontains=candidate) |
            Q(candidate__email__icontains=candidate)
        )
    
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="resumes_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    response.write('\ufeff')
    
    writer = csv.writer(response)
    writer.writerow([
        'ID', 'Кандидат', 'Email', 'Файл', 'Размер (KB)', 'Язык', 
        'Статус', 'Опыт (лет)', 'Навыки', 'Дата загрузки'
    ])
    
    for resume in resumes:
        writer.writerow([
            resume.id,
            resume.candidate.full_name,
            resume.candidate.email,
            resume.original_filename,
            f"{resume.file_size / 1024:.1f}",
            resume.get_language_display(),
            resume.get_status_display(),
            resume.experience_years or '-',
            ', '.join(resume.skills[:5]) if resume.skills else '-',
            resume.uploaded_at.strftime('%d.%m.%Y %H:%M')
        ])
    
    return response

def export_resumes_pdf(request):
    import io
    import os

    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.pdfgen import canvas

    status = request.GET.get('status')
    language = request.GET.get('language')
    candidate = request.GET.get('candidate')
    
    resumes = Resume.objects.select_related('candidate').order_by('-uploaded_at')
    
    if status:
        resumes = resumes.filter(status=status)
    if language:
        resumes = resumes.filter(language=language)
    if candidate:
        resumes = resumes.filter(
            Q(candidate__first_name__icontains=candidate) |
            Q(candidate__last_name__icontains=candidate) |
            Q(candidate__email__icontains=candidate)
        )
    
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    try:
        # Для Windows
        if os.path.exists('C:/Windows/Fonts/arial.ttf'):
            pdfmetrics.registerFont(TTFont('Arial', 'C:/Windows/Fonts/arial.ttf'))
            font_name = 'Arial'
        # Для Linux
        elif os.path.exists('/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'):
            pdfmetrics.registerFont(TTFont('DejaVu', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
            font_name = 'DejaVu'
        # Для macOS
        elif os.path.exists('/System/Library/Fonts/Arial.ttf'):
            pdfmetrics.registerFont(TTFont('Arial', '/System/Library/Fonts/Arial.ttf'))
            font_name = 'Arial'
        else:
            # Если шрифт не найден, используем встроенный (без кириллицы)
            font_name = 'Helvetica'
    except:
        font_name = 'Helvetica'
    
    # Заголовок
    p.setFont(font_name, 16)
    if font_name == 'Helvetica':
        title = "Resume Report"  # Английский заголовок если нет кириллицы
        date_label = "Generated:"
        total_label = "Total records:"
    else:
        title = "Отчет по резюме кандидатов"
        date_label = "Дата формирования:"
        total_label = "Всего записей:"
    
    p.drawString(50, height - 50, title)
    
    p.setFont(font_name, 10)
    p.drawString(50, height - 70, f"{date_label} {timezone.now().strftime('%d.%m.%Y %H:%M')}")
    p.drawString(50, height - 85, f"{total_label} {resumes.count()}")
    
    # Таблица
    y_position = height - 120
    row_height = 20
    
    # Заголовки таблицы
    p.setFont(font_name, 10)
    if font_name == 'Helvetica':
        headers = ["ID", "Candidate", "Email", "File", "Status", "Date"]
    else:
        headers = ["ID", "Кандидат", "Email", "Файл", "Статус", "Дата"]
    
    x_positions = [50, 80, 200, 350, 450, 520]
    
    # Рисуем заголовки
    for i, header in enumerate(headers):
        p.drawString(x_positions[i], y_position, header)
    
    # Линия под заголовками
    p.line(50, y_position - 5, 580, y_position - 5)
    y_position -= row_height
    
    # Данные
    p.setFont(font_name, 8)
    count = 0
    for resume in resumes:
        if y_position < 50:
            p.showPage()
            y_position = height - 50
            # Повторяем заголовки на новой странице
            p.setFont(font_name, 10)
            for i, header in enumerate(headers):
                p.drawString(x_positions[i], y_position, header)
            p.line(50, y_position - 5, 580, y_position - 5)
            y_position -= row_height
            p.setFont(font_name, 8)
        
        if font_name == 'Helvetica':
            candidate_name = transliterate_russian(resume.candidate.full_name)
            status_text = translate_status(resume.status)
        else:
            candidate_name = resume.candidate.full_name
            status_text = resume.get_status_display()
        
        data = [
            str(resume.id),
            candidate_name[:20],
            resume.candidate.email[:25],
            resume.original_filename[:20],
            status_text,
            resume.uploaded_at.strftime('%d.%m.%Y')
        ]
        
        for i, item in enumerate(data):
            try:
                p.drawString(x_positions[i], y_position, str(item))
            except:
                safe_item = ''.join(c if ord(c) < 256 else '?' for c in str(item))
                p.drawString(x_positions[i], y_position, safe_item)
        
        y_position -= row_height
        count += 1
        
        # Ограничиваем количество записей на странице
        if count >= 100:  # Максимум 100 записей
            break
    
    p.showPage()
    p.save()
    
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="resumes_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    return response

def transliterate_russian(text):
    """Простая транслитерация кириллицы в латиницу"""
    translit_dict = {
        'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e', 'ё': 'yo',
        'ж': 'zh', 'з': 'z', 'и': 'i', 'й': 'y', 'к': 'k', 'л': 'l', 'м': 'm',
        'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r', 'с': 's', 'т': 't', 'у': 'u',
        'ф': 'f', 'х': 'h', 'ц': 'ts', 'ч': 'ch', 'ш': 'sh', 'щ': 'sch',
        'ъ': '', 'ы': 'y', 'ь': '', 'э': 'e', 'ю': 'yu', 'я': 'ya',
        'А': 'A', 'Б': 'B', 'В': 'V', 'Г': 'G', 'Д': 'D', 'Е': 'E', 'Ё': 'Yo',
        'Ж': 'Zh', 'З': 'Z', 'И': 'I', 'Й': 'Y', 'К': 'K', 'Л': 'L', 'М': 'M',
        'Н': 'N', 'О': 'O', 'П': 'P', 'Р': 'R', 'С': 'S', 'Т': 'T', 'У': 'U',
        'Ф': 'F', 'Х': 'H', 'Ц': 'Ts', 'Ч': 'Ch', 'Ш': 'Sh', 'Щ': 'Sch',
        'Ъ': '', 'Ы': 'Y', 'Ь': '', 'Э': 'E', 'Ю': 'Yu', 'Я': 'Ya'
    }
    
    result = ''
    for char in text:
        result += translit_dict.get(char, char)
    return result

def translate_status(status):
    """Перевод статусов на английский"""
    status_dict = {
        'uploaded': 'Uploaded',
        'processing': 'Processing',
        'processed': 'Processed',
        'error': 'Error'
    }
    return status_dict.get(status, status)


@login_required
def resume_list(request):
    queryset = Resume.objects.select_related('candidate').order_by('-uploaded_at')
    
    # Если пользователь - кандидат, показываем только его резюме
    if request.user.is_candidate():
        try:
            candidate = Candidate.objects.get(user=request.user)
            queryset = queryset.filter(candidate=candidate)
        except Candidate.DoesNotExist:
            queryset = Resume.objects.none()
    
    status_filter = request.GET.get('status')
    if status_filter:
        queryset = queryset.filter(status=status_filter)
    
    language_filter = request.GET.get('language')
    if language_filter:
        queryset = queryset.filter(language=language_filter)
    
    candidate_filter = request.GET.get('candidate')
    if candidate_filter:
        queryset = queryset.filter(
            Q(candidate__first_name__icontains=candidate_filter) |
            Q(candidate__last_name__icontains=candidate_filter) |
            Q(candidate__email__icontains=candidate_filter)
        )
    
    resumes = queryset
    
    # Подсчитываем статистику
    stats = queryset.aggregate(
        total=Count('id'),
        uploaded=Count('id', filter=Q(status='uploaded')),
        processing=Count('id', filter=Q(status='processing')),
        processed=Count('id', filter=Q(status='processed')),
        error=Count('id', filter=Q(status='error')),
    )
    
    context = {
        'resumes': resumes,
        'total_resumes': stats['total'],
        'uploaded_resumes': stats['uploaded'],
        'processing_resumes': stats['processing'],
        'processed_resumes': stats['processed'],
        'error_resumes': stats['error'],
        'can_view_sensitive_data': can_view_sensitive_data(request.user),
        'can_manage_resumes': can_manage_applications(request.user),
        'can_export_data': can_export_data(request.user),
        'is_candidate': request.user.is_candidate(),
    }
    
    return render(request, 'resume/resume_list.html', context)

def resume_detail(request, pk):
    resume = get_object_or_404(Resume, pk=pk)
    return render(request, 'resume/resume_detail.html', {'resume': resume})
